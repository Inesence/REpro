# -*- coding: utf-8 -*-
"""
Created on Sat Feb 18 19:39:44 2023

@author: Inese
"""

import pandas as pd

from openpyxl import load_workbook
from Automated_receipt_processing import create_dataframe_from_jsons,format_df, concat_columns, sort_df

# Set the directory where the JSON files are stored
json_dir = '/path/to/json/files/'  # !!! Insert the path to your key folder here!!!!


df = create_dataframe_from_jsons(json_dir)
df=format_df(df)
concat_df=concat_columns(df, ["nosaukums","PVN_maksataja_numurs","receipt_nr"], "apraksts")
sorted_df=sort_df(concat_df)



def add_data_to_excel(df, file_name):
    # Open the Excel file
    wb = load_workbook(file_name)
    # Group the data by month
    df['month'] = pd.to_datetime(df['datums'], format='%d.%m.%Y').dt.to_period('M')
    groups = df.groupby('month')
    
    # Loop through the groups and save the data into separate sheets
    for name, group in groups:
        # Create a new sheet for the month if it doesn't already exist
        if name.strftime('%m_%y') not in wb.sheetnames:
            sheet = wb.copy_worksheet(wb['Template'])
            sheet.title = name.strftime('%m_%y')
        else:
            sheet = wb[name.strftime('%m_%y')]
        # Find the row to insert the new data
        

        # Insert the new data into the sheet in chronological order
        for i, row in group.iterrows():
            row_to_insert = 29
            while sheet.cell(row=row_to_insert, column=2).value is not None and sheet.cell(row=row_to_insert, column=2).value < row['datums']:
                row_to_insert += 1
            sheet.insert_rows(row_to_insert)
            sheet.cell(row=row_to_insert, column=2).value = row['datums']
            sheet.cell(row=row_to_insert, column=3).value = row['apraksts']
            sheet.cell(row=row_to_insert, column=4).value = row['summa']
            

    # Save the output file
    wb.save(file_name)    
    
# Instead of "2022_avanss.xlsx" add the name of your file or a path to it if not in the same folder as code
add_data_to_excel(df, "2022_avanss.xlsx")