# -*- coding: utf-8 -*-
"""
Created on Sat Feb 18 19:39:44 2023

@author: Inese
"""

import pandas as pd
import json
import os
from openpyxl import load_workbook

# Set the directory where the JSON files are stored
json_dir = '/path/to/json/files/' # !!! Insert the path to your key folder here!!!!



# Create an empty list to hold the data from each file

def create_dataframe_from_jsons(json_dir):

    key_list = []
    value_lists = []
    # Loop through each file in the directory
    for filename in os.listdir(json_dir):
        if filename.endswith('.json'):
            # Open the file and load the JSON data into a list of dictionaries
            with open(os.path.join(json_dir, filename), 'r', encoding='utf-8') as f:
                data = json.load(f)
                data = dict(sorted(data.items()))
                # Extract the keys and values
                keys = list(data.keys())
                values = list(data.values())
                # Add the keys to the key list if it's the first file
                if not key_list:
                    key_list = keys
                # Check that the keys are the same for all files
                elif keys != key_list:
                    raise ValueError(f"Keys do not match for file {filename}")
                # Add the values to the value lists
                value_lists.append(values)
    
    # Create the dataframe using the key list and value lists
    df = pd.DataFrame(value_lists, columns=key_list)
    return df

df = create_dataframe_from_jsons(json_dir)

def date_format(df):

    # convert the 'dates' column to datetime format
    df['datums'] = pd.to_datetime(df['datums'], dayfirst=True, errors='coerce')
    # convert the 'dates' column to the required format
    df['datums'] = df['datums'].dt.strftime('%d.%m.%Y')

    return df


def format_df (df):
    df['PVN_maksataja_numurs'] = df['PVN_maksataja_numurs'].apply(lambda x: 'LV' + x)
    df['summa'] = df['summa'].str.replace('.', ',')
    df=date_format(df)
    
    return df

df=format_df(df)



def concat_columns(df, columns_to_concat, new_column_name):
    concatdf = df
    concatdf['receipt_nr'] = df['receipt_nr'].apply(lambda x: 'čeks Nr. ' + x)
    concatdf[new_column_name] = df[columns_to_concat].apply(lambda x: ', '.join(x.dropna().astype(str)), axis=1)
    concatdf = concatdf.drop(columns_to_concat, axis=1)
    return concatdf

concat_df=concat_columns(df, ["nosaukums","PVN_maksataja_numurs","receipt_nr"], "apraksts")



def sort_df(df):
        # combine date and time columns into a single datetime column
    df['datetime'] = pd.to_datetime(df['datums'] + ' ' + df['laiks'],dayfirst=True)
    
    # sort by datetime column
    df = df.sort_values('datetime')
    
    # drop the datetime column (if not needed)
    #df = df.drop('datetime', axis=1)
    
    return df


sorted_df=sort_df(concat_df)


def save_data_to_excel(df, template_file, output_file):
    # Group the data by month
    df['month'] = pd.to_datetime(df['datums'], format='%d.%m.%Y').dt.to_period('M')
    groups = df.groupby('month')

    # Load the template file
    book = load_workbook(template_file)

    # Loop through the groups and save the data into separate sheets
    for name, group in groups:
        # Create a new sheet for the month
        sheet = book.copy_worksheet(book['Template'])
        sheet.title = name.strftime('%m_%y')
        k=1
        # Write the data to the sheet
        for i, row in group.iterrows():
    
            sheet.cell(row=k+28, column=1).value = k
            sheet.cell(row=k+28, column=2).value = row['datums']
            sheet.cell(row=k+28, column=3).value = row['apraksts']
            sheet.cell(row=k+28, column=4).value = row['summa']
            k=k+1

    # Save the output file
    book.save(output_file)

# Provide path to your template file if not in the same folder and a name for your output file
save_data_to_excel(sorted_df, "template.xlsx", "2022_avanss.xlsx")